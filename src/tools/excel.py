from datetime import date, datetime
from itertools import zip_longest
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _cell(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10] if isinstance(value, date) and not isinstance(value, datetime) else value.isoformat()
    return value


def _sheet_to_rows(ws) -> list[dict]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return []
    return [
        {h: _cell(v) for h, v in zip_longest(header, row, fillvalue=None) if h is not None}
        for row in rows_iter
        if any(c is not None for c in row)
    ]


def parse_xlsx(path: str, sheet: str | None = None) -> dict[str, list[dict]] | list[dict]:
    """Parse an .xlsx workbook. Returns {sheet_name: [row_dicts]} or [row_dicts] if `sheet` given."""
    if not Path(path).exists():
        raise FileNotFoundError(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet is not None:
            if sheet not in wb.sheetnames:
                raise ValueError(f"sheet {sheet!r} not in {wb.sheetnames}")
            return _sheet_to_rows(wb[sheet])
        return {name: _sheet_to_rows(wb[name]) for name in wb.sheetnames}
    finally:
        wb.close()
